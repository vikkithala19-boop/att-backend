from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import requests
from bs4 import BeautifulSoup
import time
import os

from openpyxl import Workbook, load_workbook
from datetime import datetime

from fetch_courses import fetch_courses
from fetch_attendance import fetch_attendance

app = Flask(__name__)
CORS(app)

LOGIN_URL = "https://ecampus.psgtech.ac.in/studzone"

HEADERS = {
    "User-Agent": "Mozilla/5.0"
}

# cache
CACHE = {}
CACHE_TTL = 60


# ✅ SAFE LOG FUNCTION
def log_user(roll):
    try:
        file_name = "users.xlsx"

        now = datetime.now()
        date = now.strftime("%d-%m-%Y")
        time_now = now.strftime("%H:%M:%S")

        # create file if not exists
        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.append(["Roll Number", "Date", "Time"])
            wb.save(file_name)

        wb = load_workbook(file_name)
        ws = wb.active

        ws.append([roll, date, time_now])
        wb.save(file_name)

    except Exception as e:
        print("LOG ERROR:", e)   # ⚠️ prevent crash


@app.route("/")
def home():
    return "PSG Attendance Backend Running"


def login_psg(roll, password):

    session = requests.Session()

    r = session.get(LOGIN_URL, headers=HEADERS)

    soup = BeautifulSoup(r.text, "html.parser")

    token = soup.find("input", {"name": "__RequestVerificationToken"})

    if not token:
        raise Exception("Login token not found")

    token = token["value"]

    payload = {
        "__RequestVerificationToken": token,
        "RollNo": roll,
        "Password": password
    }

    login = session.post(LOGIN_URL, data=payload, headers=HEADERS)

    if "Invalid" in login.text:
        raise Exception("Invalid credentials")

    return session


@app.route("/attendance", methods=["POST"])
def attendance():

    global CACHE

    data = request.json

    roll = data.get("id")
    password = data.get("password")

    if not roll or not password:
        return jsonify({"error": "Missing credentials"})

    # ✅ log user safely
    log_user(roll)

    # cache check
    if roll in CACHE:
        cached = CACHE[roll]
        if time.time() - cached["time"] < CACHE_TTL:
            return jsonify(cached["data"])

    try:

        session = login_psg(roll, password)

        subject_map = fetch_courses(session)

        attendance, subjects, last_updated = fetch_attendance(session, subject_map)

        bunkable = 0

        if attendance >= 75:
            total = 100
            attended = (attendance / 100) * total
            bunkable = int((attended / 0.75) - total)

        result = {
            "attendance": attendance,
            "bunkable": bunkable,
            "updated": last_updated,
            "subjects": subjects
        }

        CACHE[roll] = {
            "data": result,
            "time": time.time()
        }

        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e)})


# ✅ SAFE DOWNLOAD ROUTE
@app.route("/download")
def download_file():

    file_name = "users.xlsx"

    if not os.path.exists(file_name):
        return jsonify({"error": "No data yet"})

    return send_file(file_name, as_attachment=True)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
